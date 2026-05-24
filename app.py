from __future__ import annotations

import html
import json
import mimetypes
import os
import re
import shutil
import smtplib
import threading
import urllib.error
import urllib.request
from datetime import datetime
from email.message import EmailMessage
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = Path("G:/downloads/Creative_Request_Portal_Forsaken_Form.xlsx")
DEFAULT_GOOGLE_SHEET_ID = "1vF7H7Yp7MrHOKe4j6HRkjjYrpxTtEh5ugEQ_OpKDaYU"
DEFAULT_GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1vF7H7Yp7MrHOKe4j6HRkjjYrpxTtEh5ugEQ_OpKDaYU/edit?usp=drivesdk"
ENV_PATH = BASE_DIR / ".env"
LOCK = threading.Lock()

TRACKER_SHEET = "Request Tracker"
LISTS_SHEET = "Lists"
DEFAULT_DRIVE_FOLDER = "https://drive.google.com/drive/folders/12mxrlRV2QdA6UIu21Y7YTKCer_ntdTkY?usp=sharing"
DEFAULT_UPLOAD_FOLDER = "https://drive.google.com/drive/folders/17Elym_RLgFLL2EPOOgS2FKhwNA3ikd-f?usp=sharing"

FIELD_LABELS = {
    "member_name": "Member Name",
    "email": "Email",
    "project_name": "Project Name",
    "design_type": "Design Type",
    "description": "Design Description",
    "requested_deadline": "Requested Deadline",
    "priority": "Priority Level",
    "rush_option": "Expedited Option",
    "brand_guidelines_link": "Brand Guidelines Link",
    "reference_files_link": "Reference Files Link",
    "uploaded_files_link": "Uploaded Assets Folder Link",
    "notes": "Notes for Designer",
}

REQUIRED_FIELDS = [
    "member_name",
    "email",
    "project_name",
    "design_type",
    "description",
    "requested_deadline",
    "priority",
    "rush_option",
]


def load_env() -> dict[str, str]:
    env = dict(os.environ)
    if ENV_PATH.exists():
        for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            value = value.strip().strip('"').strip("'")
            env.setdefault(key.strip(), value)
    return env


def get_config() -> dict[str, str]:
    env = load_env()
    return {
        "workbook_path": env.get("WORKBOOK_PATH", str(DEFAULT_WORKBOOK)),
        "google_sheet_id": env.get("GOOGLE_SHEET_ID", DEFAULT_GOOGLE_SHEET_ID),
        "google_sheet_url": env.get("GOOGLE_SHEET_URL", DEFAULT_GOOGLE_SHEET_URL),
        "google_service_account_json": env.get("GOOGLE_SERVICE_ACCOUNT_JSON", ""),
        "google_apps_script_webhook_url": env.get("GOOGLE_APPS_SCRIPT_WEBHOOK_URL", ""),
        "google_apps_script_secret": env.get("GOOGLE_APPS_SCRIPT_SECRET", ""),
        "admin_email": env.get("ADMIN_EMAIL", ""),
        "smtp_host": env.get("SMTP_HOST", ""),
        "smtp_port": env.get("SMTP_PORT", "587"),
        "smtp_user": env.get("SMTP_USER", ""),
        "smtp_pass": env.get("SMTP_PASS", ""),
        "smtp_from": env.get("SMTP_FROM", env.get("SMTP_USER", "")),
        "smtp_tls": env.get("SMTP_TLS", "true").lower(),
        "company_team": env.get("COMPANY_TEAM", "Forsaken"),
        "drive_folder": env.get("DRIVE_FOLDER", DEFAULT_DRIVE_FOLDER),
        "upload_folder": env.get("UPLOAD_FOLDER", DEFAULT_UPLOAD_FOLDER),
        "host": env.get("HOST", "127.0.0.1"),
        "port": env.get("PORT", "8000"),
    }


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def read_choices() -> dict[str, list[str]]:
    config = get_config()
    workbook_path = Path(config["workbook_path"])
    choices = {
        "designTypes": [
            "AVIs",
            "Twitter Headers",
            "YouTube Headers",
            "TikTok Banners",
            "Social Media Packages",
            "3D Animations",
            "Intros or Outros",
            "Simple Editing",
            "Advanced Editing",
            "Other",
        ],
        "priorities": ["Standard", "Expedited"],
        "rushOptions": ["No Rush", "24 Hour Rush +50%", "Same Day / Emergency Rush +100%"],
    }
    if not workbook_path.exists():
        return choices

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if LISTS_SHEET not in wb.sheetnames:
            return choices
        ws = wb[LISTS_SHEET]
        choices["designTypes"] = non_empty_column(ws, 2)
        choices["priorities"] = non_empty_column(ws, 3)
        choices["rushOptions"] = non_empty_column(ws, 6)
    finally:
        wb.close()
    return choices


def non_empty_column(ws: Any, column: int) -> list[str]:
    values: list[str] = []
    for row in ws.iter_rows(min_row=2, min_col=column, max_col=column, values_only=True):
        value = cell_text(row[0])
        if value:
            values.append(value)
    return values


def parse_form(body: bytes) -> dict[str, str]:
    parsed = parse_qs(body.decode("utf-8"), keep_blank_values=True)
    return {key: values[0].strip() for key, values in parsed.items()}


def validate_form(data: dict[str, str]) -> list[str]:
    errors = []
    for field in REQUIRED_FIELDS:
        if not data.get(field):
            errors.append(f"{FIELD_LABELS[field]} is required.")
    if data.get("email") and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", data["email"]):
        errors.append("Enter a valid email address.")
    return errors


def next_request_id(ws: Any) -> str:
    highest = 0
    for row in range(2, ws.max_row + 1):
        value = cell_text(ws.cell(row=row, column=1).value)
        match = re.fullmatch(r"CRP-(\d+)", value)
        if match:
            highest = max(highest, int(match.group(1)))
    return f"CRP-{highest + 1:04d}"


def first_empty_request_row(ws: Any) -> int:
    for row in range(2, max(ws.max_row, 200) + 1):
        if not cell_text(ws.cell(row=row, column=1).value):
            return row
    return ws.max_row + 1


def rush_fee_text(rush_option: str) -> str:
    if "Same Day" in rush_option or "Emergency" in rush_option:
        return "+100% rush fee"
    if "24 Hour" in rush_option:
        return "+50% rush fee"
    return "No additional fee"


def combined_admin_notes(data: dict[str, str]) -> str:
    parts = []
    if data.get("brand_guidelines_link"):
        parts.append(f"Brand guidelines: {data['brand_guidelines_link']}")
    if data.get("reference_files_link"):
        parts.append(f"Reference files: {data['reference_files_link']}")
    if data.get("notes"):
        parts.append(f"Notes: {data['notes']}")
    if data.get("other_design_type"):
        parts.append(f"Other design type: {data['other_design_type']}")
    return "\n".join(parts)


def build_tracker_values(data: dict[str, str], request_id: str, now: datetime, config: dict[str, str], *, as_text: bool) -> list[Any]:
    uploaded_link = data.get("uploaded_files_link") or ""
    notes = combined_admin_notes(data)
    submitted = now.strftime("%Y-%m-%d %I:%M %p") if as_text else now
    deadline = data["requested_deadline"]
    return [
        request_id,
        submitted,
        data["member_name"],
        config["company_team"],
        data["email"],
        data["project_name"],
        data["design_type"],
        data["description"],
        deadline,
        data["priority"],
        data["rush_option"],
        rush_fee_text(data["rush_option"]),
        "Submitted",
        "Not Seen",
        "",
        "Pending Review",
        "",
        "",
        "",
        config["drive_folder"],
        uploaded_link,
        "",
        submitted,
        notes,
    ]


def next_google_request_id(values: list[str]) -> str:
    highest = 0
    for value in values:
        match = re.fullmatch(r"CRP-(\d+)", cell_text(value))
        if match:
            highest = max(highest, int(match.group(1)))
    return f"CRP-{highest + 1:04d}"


def save_submission_to_apps_script(data: dict[str, str], config: dict[str, str]) -> dict[str, str]:
    if not config["google_apps_script_webhook_url"]:
        raise RuntimeError(
            "Requests are set to record in Google Sheets, but the Apps Script webhook URL is not configured in Vercel."
        )
    now = datetime.now()
    payload = {
        "secret": config["google_apps_script_secret"],
        "submitted_at": now.strftime("%Y-%m-%d %I:%M %p"),
        "values": build_tracker_values(data, "", now, config, as_text=True),
        "fields": data,
        "files": json.loads(data.get("uploaded_files_json") or "[]"),
    }
    body = json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        config["google_apps_script_webhook_url"],
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=55) as response:
            response_payload = json.loads(response.read().decode("utf-8") or "{}")
    except urllib.error.HTTPError as exc:
        message = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Google Apps Script rejected the request: {message}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"Could not reach the Google Apps Script webhook: {exc.reason}") from exc

    if not response_payload.get("ok"):
        raise RuntimeError(response_payload.get("error") or "Google Apps Script did not confirm the request was saved.")

    return {
        "request_id": response_payload.get("request_id", "CRP-SAVED"),
        "submitted_at": now.strftime("%Y-%m-%d %I:%M %p"),
        "email_sent": "false",
        "storage": "apps_script",
    }


def save_submission_to_google_sheet(data: dict[str, str], config: dict[str, str]) -> dict[str, str]:
    if not config["google_service_account_json"]:
        raise RuntimeError(
            "Requests are set to record in Google Sheets, but the site still needs the Google Sheets connection added in Vercel."
        )
    try:
        import gspread
    except ImportError as exc:
        raise RuntimeError("Google Sheets support requires gspread from requirements.txt.") from exc

    credentials = json.loads(config["google_service_account_json"])
    client = gspread.service_account_from_dict(credentials)
    spreadsheet = client.open_by_key(config["google_sheet_id"])
    worksheet = spreadsheet.worksheet(TRACKER_SHEET)
    request_id = next_google_request_id(worksheet.col_values(1))
    now = datetime.now()
    values = build_tracker_values(data, request_id, now, config, as_text=True)
    worksheet.append_row(values, value_input_option="USER_ENTERED")
    return {
        "request_id": request_id,
        "submitted_at": now.strftime("%Y-%m-%d %I:%M %p"),
        "email_sent": "false",
        "storage": "google_sheet",
    }


def save_submission(data: dict[str, str]) -> dict[str, str]:
    config = get_config()
    workbook_path = Path(config["workbook_path"])
    if config["google_apps_script_webhook_url"]:
        return save_submission_to_apps_script(data, config)
    if os.environ.get("VERCEL"):
        return save_submission_to_google_sheet(data, config)
    if config["google_service_account_json"]:
        return save_submission_to_google_sheet(data, config)
    if not workbook_path.exists():
        raise FileNotFoundError(
            "Requests are set to record in Google Sheets, but the Google Sheets connection is not configured yet."
        )

    with LOCK:
        backup_dir = BASE_DIR / "backups"
        backup_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        shutil.copy2(workbook_path, backup_dir / f"{workbook_path.stem}-{timestamp}.xlsx")

        wb = load_workbook(workbook_path)
        try:
            if TRACKER_SHEET not in wb.sheetnames:
                raise ValueError(f"Sheet not found: {TRACKER_SHEET}")
            ws = wb[TRACKER_SHEET]
            request_id = next_request_id(ws)
            row = first_empty_request_row(ws)
            now = datetime.now()

            values = build_tracker_values(data, request_id, now, config, as_text=False)
            for col, value in enumerate(values, start=1):
                ws.cell(row=row, column=col, value=value)
            ws.cell(row=row, column=2).number_format = "yyyy-mm-dd h:mm AM/PM"
            ws.cell(row=row, column=23).number_format = "yyyy-mm-dd h:mm AM/PM"
            wb.save(workbook_path)
        finally:
            wb.close()

    return {
        "request_id": request_id,
        "submitted_at": datetime.now().strftime("%Y-%m-%d %I:%M %p"),
        "email_sent": "false",
    }


def send_email(data: dict[str, str], result: dict[str, str]) -> bool:
    config = get_config()
    required = ["admin_email", "smtp_host", "smtp_port", "smtp_from"]
    if any(not config[key] for key in required):
        return False

    msg = EmailMessage()
    msg["Subject"] = f"New creative request: {result['request_id']} - {data['project_name']}"
    msg["From"] = config["smtp_from"]
    msg["To"] = config["admin_email"]
    if data.get("email"):
        msg["Reply-To"] = data["email"]

    lines = [
        f"Request ID: {result['request_id']}",
        f"Submitted: {result['submitted_at']}",
        "",
    ]
    for key, label in FIELD_LABELS.items():
        if data.get(key):
            lines.append(f"{label}: {data[key]}")
    msg.set_content("\n".join(lines))

    rows = "".join(
        f"<tr><th>{html.escape(label)}</th><td>{html.escape(data.get(key, ''))}</td></tr>"
        for key, label in FIELD_LABELS.items()
        if data.get(key)
    )
    msg.add_alternative(
        f"""
        <html>
          <body>
            <h2>New creative request: {html.escape(result['request_id'])}</h2>
            <p>Submitted {html.escape(result['submitted_at'])}</p>
            <table cellpadding="8" cellspacing="0" border="1">{rows}</table>
          </body>
        </html>
        """,
        subtype="html",
    )

    port = int(config["smtp_port"])
    if port == 465:
        server: smtplib.SMTP = smtplib.SMTP_SSL(config["smtp_host"], port, timeout=20)
    else:
        server = smtplib.SMTP(config["smtp_host"], port, timeout=20)
    try:
        if config["smtp_tls"] == "true" and port != 465:
            server.starttls()
        if config["smtp_user"] and config["smtp_pass"]:
            server.login(config["smtp_user"], config["smtp_pass"])
        server.send_message(msg)
    finally:
        server.quit()
    return True


def page_template(content: str, status: str = "") -> bytes:
    choices_json = json.dumps(read_choices())
    config = get_config()
    html_doc = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Forsaken Creative Request Portal</title>
  <style>
    :root {{
      color-scheme: light;
      --ink: #151515;
      --muted: #6d6d6d;
      --line: #e4e0da;
      --field: #f8f5ef;
      --accent: #e74719;
      --accent-dark: #b93412;
      --gold: #f7b733;
      --warn: #a24112;
      --ok: #166534;
      --page: #111111;
      --panel: #ffffff;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      color: var(--ink);
      background:
        linear-gradient(90deg, rgba(10, 10, 10, .94), rgba(10, 10, 10, .56)),
        url("/static/forsaken-background.png") center top / cover fixed,
        var(--page);
    }}
    body::before {{
      content: "";
      position: fixed;
      inset: -7%;
      z-index: 0;
      pointer-events: none;
      background:
        radial-gradient(circle at 16% 24%, rgba(247, 183, 51, .30), transparent 20%),
        radial-gradient(circle at 70% 18%, rgba(231, 71, 25, .34), transparent 24%),
        radial-gradient(circle at 88% 76%, rgba(255, 255, 255, .16), transparent 18%),
        linear-gradient(120deg, rgba(231, 71, 25, .16), transparent 44%, rgba(247, 183, 51, .14));
      filter: blur(24px) saturate(1.35);
      opacity: .92;
      animation: glowShift 11s ease-in-out infinite alternate;
    }}
    @keyframes glowShift {{
      0% {{ transform: translate3d(-2%, -1%, 0) scale(1); opacity: .78; }}
      100% {{ transform: translate3d(2%, 2%, 0) scale(1.04); opacity: .96; }}
    }}
    main {{
      position: relative;
      z-index: 2;
      width: min(1180px, calc(100% - 32px));
      margin: 0 auto;
      min-height: 100vh;
      display: grid;
      grid-template-columns: minmax(280px, .75fr) minmax(0, 1.25fr);
      align-items: center;
      gap: 34px;
      padding: 34px 0;
    }}
    header {{
      color: #ffffff;
      padding: 18px 0;
    }}
    .eyebrow {{
      display: inline-block;
      color: var(--gold);
      font-size: .78rem;
      font-weight: 800;
      letter-spacing: 0;
      text-transform: uppercase;
      margin-bottom: 13px;
    }}
    h1 {{
      margin: 0;
      font-size: clamp(2.65rem, 6vw, 5.6rem);
      line-height: .9;
      text-transform: uppercase;
      max-width: 8ch;
    }}
    p {{ margin: 0; color: var(--muted); line-height: 1.5; }}
    .form-shell {{
      background: rgba(255, 255, 255, .96);
      border: 1px solid rgba(255, 255, 255, .45);
      border-radius: 8px;
      box-shadow: 0 22px 70px rgba(0, 0, 0, .36);
      overflow: hidden;
    }}
    .form-title {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 18px;
      padding: 22px 26px;
      background: #161616;
      color: #ffffff;
      border-bottom: 4px solid var(--accent);
    }}
    .form-title h2 {{
      margin: 0;
      font-size: clamp(1.25rem, 2vw, 1.8rem);
      text-transform: uppercase;
    }}
    .tag {{
      color: var(--gold);
      font-weight: 800;
      font-size: .85rem;
      white-space: nowrap;
    }}
    form {{ padding: 26px; }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 18px;
    }}
    label {{ display: grid; gap: 7px; font-weight: 700; font-size: .92rem; }}
    .hidden {{ display: none; }}
    input, select, textarea {{
      width: 100%;
      border: 1px solid #d2ccc2;
      border-radius: 6px;
      background: var(--field);
      color: var(--ink);
      font: inherit;
      padding: 11px 12px;
      min-height: 44px;
    }}
    input[type="file"] {{
      padding: 9px 12px;
      cursor: pointer;
    }}
    input:focus, select:focus, textarea:focus {{
      outline: 2px solid rgba(231, 71, 25, .3);
      border-color: var(--accent);
      background: #ffffff;
    }}
    textarea {{ min-height: 132px; resize: vertical; }}
    .full {{ grid-column: 1 / -1; }}
    .hint {{ color: var(--muted); font-size: .82rem; font-weight: 400; }}
    .upload-row {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 10px;
      align-items: end;
    }}
    .upload-link {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 44px;
      padding: 0 14px;
      border-radius: 6px;
      border: 1px solid var(--accent);
      color: #ffffff;
      background: var(--accent);
      font-weight: 700;
      text-decoration: none;
      white-space: nowrap;
    }}
    .upload-link:hover {{ background: var(--accent-dark); }}
    .actions {{
      display: flex;
      align-items: center;
      gap: 14px;
      margin-top: 24px;
      flex-wrap: wrap;
    }}
    button {{
      border: 0;
      border-radius: 6px;
      background: var(--accent);
      color: white;
      font-weight: 700;
      text-transform: uppercase;
      padding: 12px 18px;
      min-height: 44px;
      cursor: pointer;
    }}
    button:hover {{ background: var(--accent-dark); }}
    .status {{
      padding: 12px 14px;
      border-radius: 6px;
      border: 1px solid var(--line);
      margin-bottom: 18px;
      color: var(--ok);
      background: #f0fdf4;
    }}
    .status.error {{
      color: var(--warn);
      background: #fff7ed;
      border-color: #fed7aa;
    }}
    .pricing-guide {{
      border-top: 1px solid var(--line);
      background: #fbfaf7;
    }}
    .pricing-guide summary {{
      display: flex;
      justify-content: space-between;
      gap: 16px;
      padding: 20px 26px;
      cursor: pointer;
      font-weight: 800;
      text-transform: uppercase;
      color: #161616;
      list-style: none;
    }}
    .pricing-guide summary::-webkit-details-marker {{ display: none; }}
    .pricing-guide summary span {{
      color: var(--accent);
      font-size: .9rem;
      white-space: nowrap;
    }}
    .pricing-body {{
      max-height: 520px;
      overflow: auto;
      padding: 0 26px 26px;
    }}
    .pricing-grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 14px;
    }}
    .service-card {{
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #ffffff;
      padding: 16px;
    }}
    .service-card h3 {{
      margin: 0 0 10px;
      font-size: 1rem;
      text-transform: uppercase;
      color: #161616;
    }}
    .service-card h4 {{
      margin: 12px 0 8px;
      color: var(--accent-dark);
      font-size: .92rem;
    }}
    .service-card ul {{
      margin: 0;
      padding-left: 18px;
      color: #3f3f3f;
      line-height: 1.55;
      font-size: .9rem;
    }}
    .client-notice {{
      margin-top: 14px;
      border-left: 4px solid var(--accent);
      background: #fff7ed;
    }}
    @media (max-width: 900px) {{
      body {{ background-attachment: scroll; }}
      main {{
        grid-template-columns: 1fr;
        min-height: auto;
        padding: 20px 0;
      }}
      h1 {{ max-width: 11ch; }}
    }}
    @media (max-width: 760px) {{
      .grid {{ grid-template-columns: 1fr; }}
      .pricing-grid {{ grid-template-columns: 1fr; }}
      .upload-row {{ grid-template-columns: 1fr; }}
      main {{ width: min(100% - 18px, 1180px); }}
      form, .form-title {{ padding-left: 18px; padding-right: 18px; }}
      .form-title {{ display: block; }}
      .tag {{ display: block; margin-top: 8px; }}
    }}
  </style>
</head>
<body>
  <main>
    <header>
      <span class="eyebrow">The Forsaken Society</span>
      <h1>Creative Request Portal</h1>
    </header>
    <section class="form-shell">
      <div class="form-title">
        <h2>Start a Design Request</h2>
        <span class="tag">Assets upload ready</span>
      </div>
      {content}
    </section>
  </main>
  <script>
    const choices = {choices_json};
    function fillSelect(id, values, selected) {{
      const select = document.getElementById(id);
      select.innerHTML = "";
      values.forEach((value) => {{
        const option = document.createElement("option");
        option.value = value;
        option.textContent = value;
        if (value === selected) option.selected = true;
        select.appendChild(option);
      }});
    }}
    fillSelect("design_type", choices.designTypes, "AVIs");
    fillSelect("priority", choices.priorities, "Standard");
    fillSelect("rush_option", choices.rushOptions, "No Rush");

    const form = document.querySelector("form");
    const designTypeSelect = document.getElementById("design_type");
    const otherDesignWrap = document.getElementById("other_design_wrap");
    const otherDesignInput = document.getElementById("other_design_type");
    const fileInput = document.getElementById("upload_files");
    const filePayload = document.getElementById("uploaded_files_json");
    const submitButton = form.querySelector("button[type='submit']");
    const maxUploadBytes = 4 * 1024 * 1024;

    function syncOtherDesignType() {{
      const isOther = designTypeSelect.value === "Other";
      otherDesignWrap.classList.toggle("hidden", !isOther);
      otherDesignInput.required = isOther;
      if (!isOther) otherDesignInput.value = "";
    }}
    designTypeSelect.addEventListener("change", syncOtherDesignType);
    syncOtherDesignType();

    function readFileAsPayload(file) {{
      return new Promise((resolve, reject) => {{
        const reader = new FileReader();
        reader.onload = () => {{
          const dataUrl = String(reader.result || "");
          resolve({{
            name: file.name,
            type: file.type || "application/octet-stream",
            data: dataUrl.split(",", 2)[1] || ""
          }});
        }};
        reader.onerror = () => reject(reader.error || new Error("Could not read file"));
        reader.readAsDataURL(file);
      }});
    }}

    form.addEventListener("submit", async (event) => {{
      const files = Array.from(fileInput.files || []);
      const totalBytes = files.reduce((sum, file) => sum + file.size, 0);
      if (totalBytes > maxUploadBytes) {{
        event.preventDefault();
        alert("Uploads must be 4 MB total or less. For larger videos, paste a share link instead.");
        return;
      }}
      if (!files.length) return;
      event.preventDefault();
      submitButton.disabled = true;
      submitButton.textContent = "Uploading...";
      try {{
        filePayload.value = JSON.stringify(await Promise.all(files.map(readFileAsPayload)));
        form.submit();
      }} catch (error) {{
        alert("Could not prepare the upload. Please try again or paste a share link.");
        submitButton.disabled = false;
        submitButton.textContent = "Submit Request";
      }}
    }});
  </script>
</body>
</html>"""
    return html_doc.encode("utf-8")


def pricing_guide_html() -> str:
    return """
    <details class="pricing-guide">
      <summary>NAXYSTUDIOS LLC — Service Menu & Pricing Guide <span>View pricing</span></summary>
      <div class="pricing-body">
        <div class="pricing-grid">
          <section class="service-card">
            <h3>Logo Services</h3>
            <h4>Standard Logo Package — Starting at $150</h4>
            <ul>
              <li>Text Logo, Minimal Logo, Monogram Logo, Icon Logo</li>
              <li>Gaming Logo, Streetwear Logo, Business Logo, Badge Logo</li>
              <li>Includes up to 2 revisions, high-resolution exports, and transparent background version</li>
            </ul>
            <h4>Advanced Logo & Branding</h4>
            <ul>
              <li>Mascot Logo — Starting at $300+</li>
              <li>Luxury Logo — Starting at $350+</li>
              <li>3D Logo — Starting at $400+</li>
              <li>Animated Logo — Starting at $450+</li>
              <li>Metallic Logo Design — Starting at $250+</li>
              <li>Vector Conversion — Starting at $75+</li>
              <li>Logo Restoration — Starting at $100+</li>
              <li>Brand Style Guide — Starting at $250+</li>
              <li>Full Brand Identity System — Starting at $800+</li>
            </ul>
          </section>
          <section class="service-card">
            <h3>Social Media Services</h3>
            <h4>Basic Social Media Designs</h4>
            <ul>
              <li>Instagram Post — Starting at $40</li>
              <li>Facebook Banner — Starting at $60</li>
              <li>YouTube Banner — Starting at $75</li>
              <li>Twitch Banner — Starting at $75</li>
              <li>Profile Picture / AVI — Starting at $50</li>
              <li>Promotional Post — Starting at $45</li>
              <li>Story Design — Starting at $35</li>
              <li>Social Ad Graphic — Starting at $60</li>
            </ul>
            <h4>Advanced Social Media</h4>
            <ul>
              <li>Animated Posts — Starting at $150+</li>
              <li>Social Media Kits — Starting at $250+</li>
              <li>Carousel Posts — Starting at $120+</li>
              <li>Viral Marketing Graphics — Starting at $175+</li>
              <li>AI Generated Social Content — Starting at $200+</li>
              <li>Social Campaign Packages — Starting at $350+</li>
              <li>Stream Overlays — Starting at $200+</li>
              <li>Discord Branding — Starting at $150+</li>
            </ul>
          </section>
          <section class="service-card">
            <h3>Website Services</h3>
            <h4>Basic Website Services</h4>
            <ul>
              <li>Landing Page — Starting at $250</li>
              <li>One-Page Website — Starting at $400</li>
              <li>Portfolio Website — Starting at $500+</li>
              <li>Small Business Website — Starting at $600+</li>
              <li>Contact Page — Starting at $120</li>
              <li>Mobile Optimization — Starting at $100+</li>
            </ul>
            <h4>Advanced Website Services</h4>
            <ul>
              <li>E-Commerce Website — Starting at $1,200+</li>
              <li>Booking System — Starting at $400+</li>
              <li>Membership Website — Starting at $1,500+</li>
              <li>Custom UI/UX Systems — Starting at $1,000+</li>
              <li>Dashboard Design — Starting at $800+</li>
              <li>Website Redesign — Starting at $750+</li>
              <li>SEO Setup — Starting at $250+</li>
              <li>Advanced Integrations — Starting at $500+</li>
              <li>AI Features & Automation — Starting at $800+</li>
            </ul>
          </section>
          <section class="service-card">
            <h3>Flyer & Print Services</h3>
            <h4>Basic Flyer Services</h4>
            <ul>
              <li>Event Flyer — Starting at $75</li>
              <li>Club Flyer — Starting at $100</li>
              <li>Promotional Flyer — Starting at $85</li>
              <li>Sale Flyer — Starting at $75</li>
              <li>Business Flyer — Starting at $90</li>
              <li>Product Flyer — Starting at $85</li>
            </ul>
            <h4>Advanced Flyer & Print</h4>
            <ul>
              <li>Animated Flyer — Starting at $175+</li>
              <li>Motion Poster — Starting at $250+</li>
              <li>Billboard Design — Starting at $300+</li>
              <li>Print Campaign Package — Starting at $500+</li>
              <li>Menu Design — Starting at $180+</li>
              <li>Large Format Print Design — Starting at $350+</li>
            </ul>
          </section>
          <section class="service-card">
            <h3>Presentation Services</h3>
            <h4>Basic Presentation Services</h4>
            <ul>
              <li>School Presentation — Starting at $120</li>
              <li>Business Presentation — Starting at $200</li>
              <li>Slide Deck — Starting at $180</li>
              <li>Portfolio Presentation — Starting at $220</li>
            </ul>
            <h4>Advanced Presentation Services</h4>
            <ul>
              <li>Animated Presentation — Starting at $450+</li>
              <li>Investor Pitch Deck — Starting at $600+</li>
              <li>Luxury Brand Presentation — Starting at $500+</li>
              <li>Interactive Slides — Starting at $350+</li>
              <li>Motion Slide Package — Starting at $400+</li>
            </ul>
          </section>
          <section class="service-card">
            <h3>Branding Services</h3>
            <h4>Basic Branding Services</h4>
            <ul>
              <li>Social Branding — Starting at $250</li>
              <li>Basic Brand Kit — Starting at $300</li>
              <li>Color Palette Setup — Starting at $100</li>
              <li>Typography Setup — Starting at $100</li>
              <li>Profile Branding — Starting at $150</li>
            </ul>
            <h4>Advanced Branding Services</h4>
            <ul>
              <li>Full Brand Identity — Starting at $800+</li>
              <li>Merchandise Branding — Starting at $350+</li>
              <li>Apparel Branding — Starting at $300+</li>
              <li>Vehicle Branding — Starting at $450+</li>
              <li>Packaging Design — Starting at $350+</li>
              <li>Creative Direction — Starting at $100/hr</li>
              <li>Brand Strategy — Starting at $500+</li>
              <li>Marketing Campaign Identity — Starting at $700+</li>
            </ul>
          </section>
          <section class="service-card">
            <h3>Other Creative Services</h3>
            <h4>Basic Creative Services</h4>
            <ul>
              <li>Basic Animation — Starting at $150</li>
              <li>Simple Motion Graphics — Starting at $175</li>
              <li>Product Mockup — Starting at $60</li>
              <li>Apparel Mockup — Starting at $75</li>
              <li>Video Thumbnail — Starting at $50</li>
              <li>DTF Print Setup — Starting at $75</li>
              <li>Photo Editing — Starting at $40</li>
            </ul>
            <h4>Advanced Creative Services</h4>
            <ul>
              <li>Cinematic Animation — Starting at $500+</li>
              <li>AI Video Rendering — Starting at $350+</li>
              <li>Promotional Video Editing — Starting at $250+</li>
              <li>VFX / Visual Effects — Starting at $400+</li>
              <li>Clothing Design — Starting at $250+</li>
              <li>Print Coordination — Starting at $100+</li>
              <li>Creative Consulting — Starting at $75/hr</li>
              <li>AI Image Generation — Starting at $150+</li>
              <li>3D Mockups — Starting at $250+</li>
              <li>Commercial Advertisements — Starting at $800+</li>
            </ul>
          </section>
          <section class="service-card">
            <h3>Rush Fees & Additional Options</h3>
            <h4>Rush Fees</h4>
            <ul>
              <li>The base cost of all design services must be paid before any rush fee is applied</li>
              <li>Rush fees are additional charges added on top of the already-paid base service price for expedited turnaround scheduling</li>
              <li>24 Hour Rush Fee — +50% rush fee. Example: Base Logo Design Paid ($150) + 24 Hour Rush Fee (+$75) = $225</li>
              <li>Same Day / Emergency Rush Fee — +100% rush fee. Example: Base Flyer Design Paid ($75) + Same Day Rush Fee (+$75) = $150</li>
              <li>Rush scheduling depends on current workload and availability</li>
              <li>Same-day projects may require full upfront payment before work begins</li>
              <li>Rush fees are non-refundable once work has started</li>
              <li>Complex projects may still require additional charges depending on scope and revisions</li>
            </ul>
            <h4>Additional Service Options</h4>
            <ul>
              <li>Commercial Licensing — Quoted per project</li>
              <li>Source Files — Starting at $50</li>
              <li>Printing Coordination — Starting at $100+</li>
              <li>Shipping Coordination — Starting at $50+</li>
              <li>Ongoing Retainers — Starting at $150/month</li>
              <li>Monthly Creative Partnerships — Custom Pricing</li>
              <li>Brand Management — Starting at $500+/month</li>
            </ul>
          </section>
          <section class="service-card client-notice full">
            <h3>Client Notice</h3>
            <ul>
              <li>Printing, manufacturing, and shipping costs are the responsibility of the client</li>
              <li>Pricing may vary depending on complexity, revisions, licensing, and turnaround time</li>
              <li>Rush projects depend on workload availability</li>
              <li>Additional revisions may require additional fees</li>
              <li>Final files will be released after final payment has been completed</li>
            </ul>
          </section>
        </div>
      </div>
    </details>
    """


def form_html(status: str = "", error: bool = False) -> str:
    status_html = f'<div class="status{" error" if error else ""}">{html.escape(status)}</div>' if status else ""
    upload_folder = html.escape(get_config()["upload_folder"], quote=True)
    return f"""
    <form method="post" action="/submit">
      {status_html}
      <div class="grid">
        <label>Member Name
          <input name="member_name" autocomplete="name" required>
        </label>
        <label>Email
          <input name="email" type="email" autocomplete="email" required>
        </label>
        <label>Project Name
          <input name="project_name" required>
        </label>
        <label>Requested Deadline
          <input name="requested_deadline" type="date" required>
        </label>
        <label>Design Type
          <select id="design_type" name="design_type" required></select>
        </label>
        <label id="other_design_wrap" class="hidden">Tell Us What You Need
          <input id="other_design_type" name="other_design_type" placeholder="Describe the custom service request">
        </label>
        <label>Priority Level
          <select id="priority" name="priority" required></select>
        </label>
        <label>Expedited Option
          <select id="rush_option" name="rush_option" required></select>
        </label>
        <div class="full upload-row">
          <label>File or URL Link
            <input name="uploaded_files_link" type="url" placeholder="Paste Google Drive, Dropbox, Canva, or website link">
          </label>
          <label>Upload Photos or Videos
            <input id="upload_files" name="upload_files" type="file" accept="image/*,video/*" multiple>
          </label>
          <input id="uploaded_files_json" name="uploaded_files_json" type="hidden">
        </div>
        <label class="full">Design Description
          <textarea name="description" required></textarea>
        </label>
        <label class="full">Notes for Designer
          <textarea name="notes"></textarea>
          <span class="hint">Status, approval, and admin tracking fields are filled automatically in the workbook.</span>
        </label>
      </div>
      <div class="actions">
        <button type="submit">Submit Request</button>
        <p>Saved to the Request Tracker sheet.</p>
      </div>
    </form>
    {pricing_guide_html()}
    """


class RequestHandler(BaseHTTPRequestHandler):
    def log_message(self, format: str, *args: Any) -> None:
        print(f"{self.address_string()} - {format % args}")

    def send_static(self) -> None:
        raw_path = unquote(self.path.split("?", 1)[0])
        relative = raw_path.removeprefix("/static/").replace("/", os.sep)
        static_root = (BASE_DIR / "static").resolve()
        file_path = (static_root / relative).resolve()
        if not str(file_path).startswith(str(static_root)) or not file_path.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        body = file_path.read_bytes()
        content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "public, max-age=3600")
        self.end_headers()
        self.wfile.write(body)

    def send_html(self, body: bytes, status: HTTPStatus = HTTPStatus.OK) -> None:
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_json(self, payload: dict[str, Any], status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        if self.path.startswith("/static/"):
            self.send_static()
            return
        if self.path == "/health":
            config = get_config()
            self.send_json({
                "ok": True,
                "workbook_exists": Path(config["workbook_path"]).exists(),
                "google_sheet_id": config["google_sheet_id"],
                "google_sheet_url": config["google_sheet_url"],
                "google_sheet_configured": bool(config["google_service_account_json"]),
                "apps_script_webhook_configured": bool(config["google_apps_script_webhook_url"]),
                "email_configured": bool(config["admin_email"] and config["smtp_host"] and config["smtp_from"]),
            })
            return
        self.send_html(page_template(form_html()))

    def do_POST(self) -> None:
        if self.path != "/submit":
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        length = int(self.headers.get("Content-Length", "0"))
        data = parse_form(self.rfile.read(length))
        errors = validate_form(data)
        if errors:
            self.send_html(page_template(form_html(" ".join(errors), error=True)), HTTPStatus.BAD_REQUEST)
            return
        try:
            result = save_submission(data)
            try:
                email_sent = send_email(data, result)
                suffix = " Email alert sent." if email_sent else " Email not sent yet because SMTP is not configured."
            except Exception as email_error:
                suffix = f" Email could not be sent: {email_error}"
            self.send_html(page_template(form_html(f"{result['request_id']} saved successfully.{suffix}")))
        except Exception as exc:
            self.send_html(page_template(form_html(str(exc), error=True)), HTTPStatus.INTERNAL_SERVER_ERROR)


def main() -> None:
    config = get_config()
    server = ThreadingHTTPServer((config["host"], int(config["port"])), RequestHandler)
    print(f"Creative request portal running at http://{config['host']}:{config['port']}")
    print(f"Workbook: {config['workbook_path']}")
    server.serve_forever()


handler = RequestHandler


if __name__ == "__main__":
    main()
